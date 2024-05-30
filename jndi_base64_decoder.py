import openpyxl
import base64
import sys
import argparse

parser=argparse.ArgumentParser(
    description=''' An ok script for decoding and deobfuscating base64-encoded payloads, read from an xlsx file.
    IMPORTANT:
    - Install openpyxl first before using this script. How? Use "pip3 install openpyxl" in Powershell
    - The script and the xlsx file must be in the same directory.
    
    Usage: python3 jndi_base64_decoder.py [name].xlsx ''')
parser.add_argument("filename")
args=parser.parse_args()

def deobfuscate(encoded):

    decoded = [str(base64.b64decode(x)) for x in encoded]
    for s in decoded:
        s = s.split("$")
        s = [x[x.find("-")+1:x.find("-")+2] for x in s if x]
        print("".join(s))

workbook = sys.argv[1]

# Define variable to load the dataframe
dt = openpyxl.load_workbook(workbook)  # ini yg diganti
 
# Define variable to read sheet
dt1 = dt.active
 
encoded = []

# Iterate the loop to read the cell values
for row in range(0, dt1.max_row):
    for col in dt1.iter_cols(1, dt1.max_column):
        if col[row].value and col[row].value !="attackcontext":
            encoded.append(col[row].value)


print(deobfuscate(encoded))
